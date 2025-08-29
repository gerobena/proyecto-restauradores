import io
import streamlit as st
import pandas as pd
import plotly.express as px
from pathlib import Path
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from streamlit_option_menu import option_menu

st.set_page_config(page_title="Dashboard Inventario", layout="wide")
st.title("PANEL DE CONTROL - SERVICENTRO RESTAURADORES")

APP_DIR = Path(__file__).resolve().parent
ROOT_DIR = APP_DIR.parent
DATA_PATH = ROOT_DIR / "data" / "summary.parquet"
LOGO_PATH = APP_DIR / "logo2.png"

@st.cache_data
def load_data(path: Path) -> pd.DataFrame:
    if not path.exists():
        st.warning("⚠️ No se encontró `data/summary.parquet`. Ejecuta el notebook o el pipeline.")
        return pd.DataFrame()
    return pd.read_parquet(path)

df = load_data(DATA_PATH)

if df.empty:
    st.info("Aún no hay datos procesados para mostrar.")
    st.stop()  # <- evita que el resto del script intente usar 'df'

# 3. Selector de División principal
# Dentro del sidebar:
with st.sidebar:
    
    # 3 columnas para centrar el logo
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        if LOGO_PATH.exists():
            st.image(str(LOGO_PATH), width="stretch")
        else:
            st.empty()
    
    seleccion_div = option_menu(
        menu_title=None,                    # sin título encima
        options=["ALMACEN", "TALLER"],      # tus dos opciones
        icons=["house", "gear"],            # íconos (de Bootstrap Icons)
        menu_icon="list",                   # ícono general (puedes cambiarlo)
        default_index=0,                    # empieza seleccionado “ALMACEN”
        orientation="vertical",             # menú en vertical
        styles={                            # personalización de colores
            "container": {"padding": "0!important", "background-color": "#111827"},
            "nav-link": {
                "font-size": "16px",
                "text-align": "left",
                "margin": "0px",
                "color": "#ffffff",
                "padding": "8px 16px"
            },
            "nav-link-selected": {
                "background-color": "#FF7F0E",
                "color": "#ffffff"
            },
            "icon": {"color": "#ffffff", "font-size": "18px"}, 
        }
    )

# 4. Filtros
# Filtrar el DataFrame por la división elegida
# Ajusta el .str.capitalize() si tus valores vienen en mayúsculas o minúsculas distintas
df = df[df["COMERCIO"] == seleccion_div]

# Solo marcas de la división elegida ———
marcas = ["Todas"] + sorted(df["MARCA"].unique().tolist())

# Menu para filtrar
categorias = ["Todas"] + sorted(df["CATEGORIA"].unique().tolist())
subcats    = ["Todas"] + sorted(df["SUBCATEGORIA"].unique().tolist())
marcas     = ["Todas"] + sorted(df["MARCA"].unique().tolist())

col1, col2, col3 = st.columns(3)
with col1:
    sel_cat = st.selectbox("Categoría", categorias)
with col2:
    sel_sub = st.selectbox("Subcategoría", subcats)
with col3:
    sel_mar = st.selectbox("Marca", marcas)

# Aplica filtros al DataFrame
df_filtrado = df.copy()
if sel_cat != "Todas":
    df_filtrado = df_filtrado[df_filtrado["CATEGORIA"] == sel_cat]
if sel_sub != "Todas":
    df_filtrado = df_filtrado[df_filtrado["SUBCATEGORIA"] == sel_sub]
if sel_mar != "Todas":
    df_filtrado = df_filtrado[df_filtrado["MARCA"] == sel_mar]

# 5. Cálculo de KPIs (usa round() en lugar de .round())
if df_filtrado.empty:
    st.warning("No hay datos para los filtros seleccionados.")
    st.stop()

avg_dsi        = float(df_filtrado["PROMEDIO_DIAS_VENTAS"].mean())
margin_pct     = float(df_filtrado["margen_bruto_%"].mean())          # ya es porcentaje (0–100)
avg_beneficio  = float(df_filtrado["beneficio_promedio_por_producto"].mean())
total_skus     = int(df_filtrado["CODIGO"].nunique())

# Mostrar KPIs (formato limpio)
k1, k2, k3, k4 = st.columns(4)
k1.metric("Días Promedio",            f"{avg_dsi:,.1f}")      # 82.5
k2.metric("Margen Bruto %",           f"{margin_pct:,.1f}%")  # 43.6%
k3.metric("Utilidad Neta Promedio",   f"${avg_beneficio:,.1f}")  # $43.3
k4.metric("Total SKUs",               f"{total_skus:,}")      # 12 (con separador de miles si aplica)

# 6 GRAFICOS
# ——— Gráfico interactivo de Tiempo Promedio de Ventas ———

# Definimos el orden lógico de las categorías y sus etiquetas legibles
categories     = ['Estrella', 'Optimo', 'Tomar Acción', 'Tomar Acción Urgente']
display_labels = ['0-60 días', '61-90 días', '91-180 días', '>180 días']

# Contamos los productos en cada bucket de tiempo y calculamos %
#df_filtrado_grafico = df_filtrado[df_filtrado["STOCK"] > 0]
df_filtrado_grafico = df_filtrado.copy()

counts      = df_filtrado_grafico["TIEMPO_PROMEDIO_VENTAS"] \
                .value_counts() \
                .reindex(categories, fill_value=0)
total       = counts.sum()
percentages = (counts / total * 100).round(1)

# Creamos un DataFrame para Plotly
data_plotly = (
    pd.DataFrame({
        "Rango de Días": display_labels,
        "Cantidad": counts.values,
        "Porcentaje": percentages.values
    })
)

# Genera el bar chart interactivo
fig1 = px.bar(
    data_plotly,
    x="Rango de Días",
    y="Cantidad",
    text="Porcentaje",        # muestra el % encima de cada barra
    labels={                   # etiquetas de ejes
        "Cantidad": "Cantidad de Productos"
    },
    title=f"Tiempo Promedio de Ventas - {seleccion_div}",
    category_orders={"Rango de Días": display_labels},  # conserva el orden
    color_discrete_sequence=["#FF7F0E"],   # <— naranja
)

# Ajustes estéticos
fig1.update_traces(
    marker_color="#FF7F0E",
    marker_line_width=0,      # sin borde en las barras
    hovertemplate=
        "<b>%{x}</b><br>" +
        "Cantidad: %{y}<br>" +
        "Porcentaje: %{text}%" +
        "<extra></extra>",     # quita el recuadro extra de trace
    
    textfont=dict(size=14, color="black"),   # tamaño y color del texto
    textposition="inside",                   # coloca el texto fuera de la barra
    texttemplate="%{text:.1f}%"
)
fig1.update_layout(
    uniformtext_minsize=12,   # tamaño mínimo del texto
    uniformtext_mode='hide',  # esconde el texto si no cabe
    yaxis=dict(tickformat="d"),  # ejes con números enteros
    title=dict(
        text=fig1.layout.title.text,
        x=0,
        xanchor="left"
    ),
    margin=dict(t=60, b=40)
)

# 7. GRAFICO Margen Bruto

# Definimos el orden interno de las categorías y etiquetas legibles
categories     = ['Perdida', 'Beneficio Bajo', 'Beneficio Medio', 'Beneficio Alto']
display_labels = ['<0 $', '0-10 $', '10-100 $', '>100 $']

# Contamos y calculamos %
counts      = df_filtrado_grafico["metrica_margen_bruto"] \
                  .value_counts() \
                  .reindex(categories, fill_value=0)
total       = counts.sum()
percentages = (counts / total * 100).round(1)

# Preparamos un DataFrame para Plotly
data_plotly = pd.DataFrame({
    "Rango de Utilidad Neta": display_labels,
    "Cantidad": counts.values,
    "Porcentaje": percentages.values
})

# Creamos el bar chart
fig2 = px.bar(
    data_plotly,
    x="Rango de Utilidad Neta",
    y="Cantidad",
    text="Porcentaje",       # muestra el % en la barra
    category_orders={"Rango de Utilidad Neta": display_labels},
    title=f"Utilidad Neta por Producto - {seleccion_div}",
    color_discrete_sequence=["#FF7F0E"],   # <— naranja
)

# Ajustes estéticos
fig2.update_traces(\
    marker_color="#FF7F0E",
    marker_line_width=0,      # sin borde en las barras
    hovertemplate=
        "<b>%{x}</b><br>" +
        "Cantidad: %{y}<br>" +
        "Porcentaje: %{text}%" +
        "<extra></extra>",     # quita el recuadro extra de trace
    
    textfont=dict(size=14, color="black"),   # tamaño y color del texto
    textposition="inside",                   # coloca el texto fuera de la barra
    texttemplate="%{text:.1f}%"
)
fig2.update_layout(
    uniformtext_minsize=12,
    uniformtext_mode='hide',
    yaxis=dict(tickformat="d"),
    title=dict(
        text=fig2.layout.title.text,
        x=0,
        xanchor="left"
    ),
    margin=dict(t=60, b=40)
)

# 7. Layout de dos columnas para los dos gráficos
c1, c2 = st.columns(2)
with c1:
    # Render en Streamlit
    st.plotly_chart(fig1, use_container_width=True)

    # 1. Filtrar los productos > 180 días y stock existente
    df_descarga = df_filtrado[
        (df_filtrado["TIEMPO_PROMEDIO_VENTAS"] == "Tomar Acción Urgente") & (df_filtrado["STOCK"] > 0)
        ][["CODIGO", "PRODUCTO", "PROMEDIO_DIAS_VENTAS", "beneficio_promedio_por_producto", "STOCK", "unidades_vendidas_totales"]]
    
    # 1.1 Ordenar de mayor a menor según PROMEDIO_DIAS_VENTAS
    df_descarga = df_descarga.sort_values("PROMEDIO_DIAS_VENTAS", ascending=False)

    # 1.2 Renombrar columnas para el Excel
    df_descarga = df_descarga.rename(columns={
        "PROMEDIO_DIAS_VENTAS": "DIAS PROMEDIO VENTAS",
        "beneficio_promedio_por_producto": "UTILIDAD NETA PROMEDIO [$]",
        "unidades_vendidas_totales": "UNIDADES VENDIDAS"
    })

    # 2. Crear un Excel en memoria y dar formato a las columnas
    towrite = io.BytesIO()
    with pd.ExcelWriter(towrite, engine="openpyxl") as writer:
        df_descarga.to_excel(writer, index=False, sheet_name=">180_dias")
        wb = writer.book
        ws = writer.sheets[">180_dias"]

        # Columna C: PROMEDIO_DIAS_VENTAS → enteros (formato "0")
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '0'

        # Columna D: beneficio_promedio_por_producto → moneda con dos decimales
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # "$#,##0.00"
    
        # Auto-ajustar anchos de columna
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            # +2 para un pequeño margen
            ws.column_dimensions[column_letter].width = max_length + 2

    # 3. Preparar buffer y botón
    towrite.seek(0)
    st.download_button(
        label="📥 Descargar productos >180 días",
        data=towrite,
        file_name="productos_mas_180_dias.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # ——————————————————————————
    # Nuevo botón: descargar productos 0-60 días
    # 1. Filtrar “Estrella” y stock>0
    df_download_0_60 = df_filtrado[
        (df_filtrado["TIEMPO_PROMEDIO_VENTAS"] == "Estrella") &
        (df_filtrado["STOCK"] < 2)
    ][["CODIGO", "PRODUCTO", "PROMEDIO_DIAS_VENTAS", "beneficio_promedio_por_producto", "margen_bruto_%", "STOCK", "unidades_vendidas_totales"]]

    # 1.2 Ajustar margen: dividir por 100 para volverlo fracción
    df_download_0_60["margen_bruto_%"] = df_download_0_60["margen_bruto_%"] / 100

    # 1.1 Renombrar columnas
    df_download_0_60 = df_download_0_60.rename(columns={
        "PROMEDIO_DIAS_VENTAS": "DIAS PROMEDIO VENTAS",
        "beneficio_promedio_por_producto": "UTILIDAD NETA PROMEDIO [$]",
        "margen_bruto_%": "MARGEN BRUTO [%]",
        "unidades_vendidas_totales": "UNIDADES VENDIDAS"
    })

    # 2. Crear Excel en memoria y aplicar formatos
    towrite0 = io.BytesIO()

    with pd.ExcelWriter(towrite0, engine="openpyxl") as writer0:
        df_download_0_60.to_excel(writer0, index=False, sheet_name="0_60_dias")
        ws0 = writer0.sheets["0_60_dias"]

        # Formateos

        # Columna C: DIAS PROMEDIO VENTAS → enteros
        for row in ws0.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '0'

        # Columna D: UTILIDAD NETA PROMEDIO [$] → moneda con dos decimales
        for row in ws0.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        # Columna E: MARGEN BRUTO [%] → porcentaje sin decimales / Para decimales usar numbers.FORMAT_PERCENTAGE_00 
        for row in ws0.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = numbers.FORMAT_PERCENTAGE

        # Auto-ajustar anchos de columna
        for col_cells in ws0.columns:
            max_length = max(len(str(cell.value)) for cell in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws0.column_dimensions[col_letter].width = max_length + 2

    towrite0.seek(0)

    # 3. Botón de descarga para 0-60 días
    st.download_button(
        label="📥 Descargar productos Estrella Faltantes",
        data=towrite0,
        file_name="productos_estrella_sin_stock.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

with c2:
    # Render en Streamlit
    st.plotly_chart(fig2, use_container_width=True)
    
    # 1. Filtrar productos con beneficio < 0$ y con stock existente
    df_descarga2 = df_filtrado[
        (df_filtrado["metrica_margen_bruto"] == "Perdida") & (df_filtrado["STOCK"]>0)
    ][["CODIGO", "PRODUCTO", "margen_bruto_%", "beneficio_promedio_por_producto", "STOCK", "unidades_vendidas_totales"]]

    # 1.1 Ordenar de menor a mayor por beneficio_promedio_por_producto
    df_descarga2 = df_descarga2.sort_values("beneficio_promedio_por_producto", ascending=True)

    # 1.2 Ajustar margen: dividir por 100 para volverlo fracción
    df_descarga2["margen_bruto_%"] = df_descarga2["margen_bruto_%"] / 100

    # 1.3 Renombrar columnas para el Excel
    df_descarga2 = df_descarga2.rename(columns={
        "margen_bruto_%": "MARGEN BRUTO [%]",
        "beneficio_promedio_por_producto": "UTILIDAD NETA PROMEDIO [$]",
        "unidades_vendidas_totales": "UNIDADES VENDIDAS"
    })

    # 2. Crear el Excel en memoria y aplicar formatos
    towrite2 = io.BytesIO()
    with pd.ExcelWriter(towrite2, engine="openpyxl") as writer2:
        df_descarga2.to_excel(writer2, index=False, sheet_name="Perdida")
        ws2 = writer2.sheets["Perdida"]

        # Formateos
        # Columna C: MARGEN BRUTO [%] → porcentaje sin decimales / Para decimales usar numbers.FORMAT_PERCENTAGE_00 
        for row in ws2.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = numbers.FORMAT_PERCENTAGE

        # Columna D: UTILIDAD NETA PROMEDIO [$] → $ con dos decimales
        for row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        # Auto-ajustar anchos de columna
        for column_cells in ws2.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            ws2.column_dimensions[column_letter].width = max_length + 2

    towrite2.seek(0)

    # 3. Botón de descarga
    st.download_button(
        label="📥 Descargar Productos con Perdida",
        data=towrite2,
        file_name="productos_perdida.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# 8. Top 5 slow movers
top_slow = (
    df_filtrado
    # 8.1 Filtrar solo productos con stock mayor a 0
    .loc[lambda d: d["STOCK"] > 0]
    # 8.2 Ordenar de mayor a menor por DSI
    .sort_values("PROMEDIO_DIAS_VENTAS", ascending=False)
    # 8.3 Tomar los primeros 5
    .head(5)
    [["PRODUCTO", "MARCA", "PROMEDIO_DIAS_VENTAS", "beneficio_promedio_por_producto", "STOCK" ]]
    .rename(columns={"PROMEDIO_DIAS_VENTAS":"DIAS PROMEDIO", "beneficio_promedio_por_producto":"UTILIDAD NETA PROMEDIO"})
)

# Creamos una copia para formatear sin alterar los datos originales
top_slow_display = top_slow.copy()

# — Quitar decimales en DSI —
top_slow_display["DIAS PROMEDIO"] = top_slow_display["DIAS PROMEDIO"].round(0).astype(int)

# — Quitar decimales en STOCK —
top_slow_display["STOCK"] = top_slow_display["STOCK"].round(0).astype(int)

# — Formatear el beneficio con signo $ y dos decimales —
top_slow_display["UTILIDAD NETA PROMEDIO"] = top_slow_display["UTILIDAD NETA PROMEDIO"] \
    .apply(lambda x: f"${x:,.2f}")

st.subheader("TOP 5 DE ITEMS CON MOVIMIENTO LENTO")
st.table(top_slow_display)